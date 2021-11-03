import sys
import openpyxl
from PyQt5.QtWidgets import QPushButton, QProgressDialog, QSizePolicy, QVBoxLayout, QStackedWidget, QApplication, QMainWindow, QAction, QGridLayout, QWidget, QTableWidget, QTableWidgetItem, QMessageBox, QFileDialog, QProgressBar
from pdf_parser import pdf_MO
from process_xml import egrn_MO
from web_parser import get_data_from_page
from xlsxwriter.workbook import Workbook
from os import rename, path
from PyQt5.QtCore import QSize



class MainWindow(QMainWindow):
    # Переопределяем конструктор класса
    def __init__(self):
        # Обязательно нужно вызвать метод супер класса
        super().__init__()
        self.setMinimumSize(QSize(700, 600))  # Устанавливаем размеры
        self.setWindowTitle("Получение информации о недвижимости")  # Устанавливаем заголовок окна

        central_widget = QWidget(self) # Создаём центральный виджет
        self.setCentralWidget(central_widget)

        self.btn1 = QPushButton("Запуск обработки XML файлов")
        self.btn2 = QPushButton("Запуск обработки PDF файлов")
        self.btn3 = QPushButton("Получение информации с Web-портала")

        self.btn1.setSizePolicy(QSizePolicy.Preferred, QSizePolicy.Expanding)
        self.btn2.setSizePolicy(QSizePolicy.Preferred, QSizePolicy.Expanding)
        self.btn3.setSizePolicy(QSizePolicy.Preferred, QSizePolicy.Expanding)

        layout = QVBoxLayout()
        central_widget.setLayout(layout)

        layout.addWidget(self.btn1)
        layout.addWidget(self.btn2)
        layout.addWidget(self.btn3)

        self.btn1.clicked.connect(self.gotoxml)
        self.btn2.clicked.connect(self.gotopdf)
        self.btn3.clicked.connect(self.gotoweb)

    def gotopdf(self):
        mpdf = MainPdf()
        widget.addWidget(mpdf)
        widget.setCurrentIndex(widget.currentIndex()+1)

    def gotoxml(self):
        mxml = MainXml()
        widget.addWidget(mxml)
        widget.setCurrentIndex(widget.currentIndex() + 1)

    def gotoweb(self):
        m_web = MainWeb()
        widget.addWidget(m_web)
        widget.setCurrentIndex(widget.currentIndex() + 1)


# Наследуемся от QMainWindow
class MainPdf(QMainWindow):
#     # Переопределяем конструктор класса
    def __init__(self):
#         # Обязательно нужно вызвать метод супер класса
        super().__init__()
        self.setMinimumSize(QSize(1400, 600))  # Устанавливаем размеры
        self.setWindowTitle("Получение информации из PDF файла")  # Устанавливаем заголовок окна

        central_widget = QWidget(self) # Создаём центральный виджет
        self.setCentralWidget(central_widget)
        self.createtoolbar()
        self.createMenuBar()


        layout = QGridLayout()
        central_widget.setLayout(layout)

        self.table = QTableWidget()  # Создаём таблицу
        self.table.setGeometry(10, 10, 1880, 900)

        self.table.setColumnCount(22)  # Устанавливаем три колонки

        self.table.setHorizontalHeaderLabels(['№', 'Объект права', 'Адрес (местоположение) объекта права', 'Общая площадь/протяженость', 'Кадастровый номер',
     'Вид зарегистрированного права', 'Номер записи в ЕГРН', 'Дата внесения записи о регистрации в ЕГРН',
     'Сведения о документе котором удостоверена гос регистрация', 'Актуальные выписки из ЕГРН', 'Документы основания для регистрации права', 'Разрешенное использование (ВРИ)',
     'ВРИ по документам', 'Категория земель', 'Срок на который установлено обременение', 'Реквизиты обязательства', 'Лицо, в пользу которого установлено ограничение прав и обременение объекта недвижимости',
     'Наличие незасвидетельствованных данных', '               Прочие ограничения ЗУ                     ', 'Тип',
    'Кадастровые номера расположенных в пределах земельного участка объектов недвижимости\n/Кадастровые номера иных объектов недвижимости, в пределах которых расположен объект недвижимости',
     'Кдастровая стоимость'])

        self.table.resizeColumnsToContents()
        layout.addWidget(self.table, 0, 0)


    def createtoolbar(self):
        getAction_1 = QAction("Загрузить выписки", self)
        getAction_1.triggered.connect(self.get_file_name)

        getAction_2 = QAction("Выгрузить в Excel", self)
        getAction_2.triggered.connect(self.fileSave)

        getAction_3 = QAction("Смена имени файлов", self)
        getAction_3.triggered.connect(self.re_name)

        getAction_4 = QAction("Очистить таблицу", self)
        getAction_4.triggered.connect(self.clear_table)

        self.toolbar = self.addToolBar('Загрузить выписки')
        self.toolbar.addAction(getAction_1)
        self.toolbar.addAction(getAction_2)
        self.toolbar.addAction(getAction_3)
        self.toolbar.addAction(getAction_4)


    def createMenuBar(self):
        self.menuBar = self.menuBar()
        fileMenu = self.menuBar.addMenu("Меню")
        back_action = QAction("Назад", self)
        back_action.triggered.connect(self.back_menu)
        fileMenu.addAction(back_action)


    def get_file_name(self):
        self.file_name = QFileDialog.getOpenFileNames(self, 'Открыть файл', '', "*.pdf")

        if not self.file_name[0]:
            return
        else:

            self.process = []
            n = 0

            dialog, bar = self.progdialog(0)
            bar.setValue(0)
            bar.setMaximum(100)
            self.prgs = 0
            for path in self.file_name[0]:
                n += 1
                self.process.append([n] + pdf_MO(path))
                self.prgs = (n // len(self.file_name[0])) * 100
                bar.setValue(self.prgs)

            self.table.setRowCount(n)
            numb_row = len(self.process)
            for row in range(numb_row):
                for column in range(22):
                    self.table.setItem(row, column, QTableWidgetItem(str(self.process[row][column])))
            self.table.resizeRowsToContents()


    def fileSave(self):
        self.fileName = QFileDialog.getSaveFileName(self, "Сохранить файл", "", "*.xlsx")

        if not self.fileName[0]:
            return
        else:

            list = []
            model = self.table.model()
            for _row in range(model.rowCount()):
                r = []
                for _column in range(model.columnCount()):
                    r.append("{}".format(model.index(_row, _column).data() or ""))
                list.append(r)

            workbook = Workbook(self.fileName[0])
            worksheet = workbook.add_worksheet()

            _header_name = ['№', 'Объект права', 'Адрес (местоположение) объекта права', 'Общая площадь/протяженость', 'Кадастровый номер',
                       'Вид зарегистрированного права', 'Номер записи в ЕГРН', 'Дата внесения записи о регистрации в ЕГРН', 'Сведения о документе котором удостоверена гос регистрация',
                       'Актуальные выписки из ЕГРН', 'Документы основания для регистрации права', 'Разрешенное использование (ВРИ)', 'ВРИ по документам',
                       'Категория земель', 'Срок на который установлено обременение', 'Реквизиты обязательства', 'Лицо, в пользу которого установлено ограничение прав и обременение объекта недвижимости',
                       'Наличие незасвидетельствованных данных', 'Прочие ограничения ЗУ', 'Тип', 'Кадастровые номера расположенных в пределах земельного участка объектов недвижимости/Кадастровые номера иных объектов недвижимости, в пределах которых расположен объект недвижимости',
                       'Кдастровая стоимость']

            for r, row in enumerate(list):
                for c, col in enumerate(row):
                    worksheet.write(0, c, _header_name[c])
                    worksheet.write(r+1, c, col)
            workbook.close()
            QMessageBox.information(self, "Success!", f"Данные сохранены в файле: \n{self.fileName[0]}")

    def re_name(self):
        self.file_name = QFileDialog.getOpenFileNames(self, 'Открыть файл', '', "*.pdf")
        self.file_name_next = self.file_name[0]

        if not self.file_name_next:
            return
        else:

            self.list = []
            model = self.table.model()
            for _row in range(model.rowCount()):
                self.list.append("{}".format(model.index(_row, 4).data() or ""))

            for files in range(len(self.file_name_next)):
                if self.file_name_next[files] == (path.dirname(self.file_name_next[files])+"/"+self.list[files].replace(":","_")+".pdf"):
                    next
                else:
                    rename(self.file_name_next[files], str(path.dirname(self.file_name_next[files])+"/"+self.list[files].replace(":","_")+".pdf"))

    def back_menu(self):
        MW = MainWindow()
        widget.addWidget(MW)
        widget.setCurrentIndex(widget.currentIndex() + 1)


    def progdialog(self, progress):
        dialog = QProgressDialog()
        dialog.setWindowTitle("Подождите идет загрузка")
        dialog.setLabelText("Подождите идет загрузка")
        bar = QProgressBar(dialog)
        bar.setTextVisible(True)
        bar.setValue(progress)
        dialog.setBar(bar)
        dialog.setMinimumWidth(700)
        dialog.show()
        return dialog, bar

    def clear_table(self):
        self.table.setRowCount(0)


class MainXml(QMainWindow):
#     # Переопределяем конструктор класса
    def __init__(self):
#         # Обязательно нужно вызвать метод супер класса
        super().__init__()
        self.setMinimumSize(QSize(1400, 600))  # Устанавливаем размеры
        self.setWindowTitle("Получение информации из XML файла")  # Устанавливаем заголовок окна

        central_widget = QWidget(self) # Создаём центральный виджет
        self.setCentralWidget(central_widget)
        self.createtoolbar()
        self.createMenuBar()
        layout = QGridLayout()
        central_widget.setLayout(layout)

        self.table = QTableWidget()  # Создаём таблицу
        self.table.setGeometry(10, 10, 1880, 900)

        self.table.setColumnCount(22)  # Устанавливаем три колонки


        # Устанавливаем заголовки таблицы
        self.table.setHorizontalHeaderLabels(['№', 'Объект права', 'Адрес (местоположение) объекта права', 'Общая площадь/протяженость', 'Кадастровый номер',
                       'Вид зарегистрированного права', 'Номер записи в ЕГРН', 'Дата внесения записи о регистрации в ЕГРН', 'Сведения о документе котором удостоверена гос регистрация',
                       'Актуальные выписки из ЕГРН', 'Документы основания для регистрации права', 'Разрешенное использование (ВРИ)', 'ВРИ по документам',
                       'Категория земель', 'Срок на который установлено обременение', 'Реквизиты обязательства', 'Лицо, в пользу которого установлено ограничение прав и обременение объекта недвижимости',
                       'Наличие незасвидетельствованных данных',              'Прочие ограничения ЗУ               ', 'Тип',
                        'Кадастровые номера расположенных в пределах земельного участка объектов недвижимости\n/Кадастровые номера иных объектов недвижимости, в пределах которых расположен объект недвижимости',
                       'Кдастровая стоимость'])

        self.table.resizeColumnsToContents()
        layout.addWidget(self.table, 0, 0)


    def createtoolbar(self):
        getAction_1 = QAction("Загрузить выписки", self)
        getAction_1.triggered.connect(self.get_file_name)

        getAction_2 = QAction("Выгрузить в Excel", self)
        getAction_2.triggered.connect(self.fileSave)

        getAction_3 = QAction("Смена имени файлов", self)
        getAction_3.triggered.connect(self.re_name)

        getAction_4 = QAction("Очистить таблицу", self)
        getAction_4.triggered.connect(self.clear_table)

        self.toolbar = self.addToolBar('Загрузить выписки')
        self.toolbar.addAction(getAction_1)
        self.toolbar.addAction(getAction_2)
        self.toolbar.addAction(getAction_3)
        self.toolbar.addAction(getAction_4)


    def createMenuBar(self):
        self.menuBar = self.menuBar()
        fileMenu = self.menuBar.addMenu("Меню")
        back_action = QAction("Назад", self)
        back_action.triggered.connect(self.back_menu)
        fileMenu.addAction(back_action)

    def get_file_name(self):
        self.file_name = QFileDialog.getOpenFileNames(self, 'Открыть файл', '', "*.xml")

        if not self.file_name[0]:
            return
        else:

            self.process = []
            n = 0
            for path in self.file_name[0]:
                n += 1
                self.process.append([n] + egrn_MO(path))

            self.table.setRowCount(len(self.process))

            numb_row = len(self.process)

            for row in range(numb_row):
                for column in range(22):
                    self.table.setItem(row, column, QTableWidgetItem(str(self.process[row][column])))

            self.table.resizeRowsToContents()


    def fileSave(self):
        self.fileName = QFileDialog.getSaveFileName(self, "Сохранить файл", ".", "*.xlsx")

        if not self.fileName[0]:
            return
        else:

            list = []
            model = self.table.model()
            for _row in range(model.rowCount()):
                r = []
                for _column in range(model.columnCount()):
                    r.append("{}".format(model.index(_row, _column).data() or ""))
                list.append(r)


            workbook = Workbook(self.fileName[0])
            worksheet = workbook.add_worksheet()

            _header_name = ['№', 'Объект права', 'Адрес (местоположение) объекта права', 'Общая площадь/протяженость', 'Кадастровый номер',
                       'Вид зарегистрированного права', 'Номер записи в ЕГРН', 'Дата внесения записи о регистрации в ЕГРН', 'Сведения о документе котором удостоверена гос регистрация',
                       'Актуальные выписки из ЕГРН', 'Документы основания для регистрации права', 'Разрешенное использование (ВРИ)', 'ВРИ по документам',
                       'Категория земель', 'Срок на который установлено обременение', 'Реквизиты обязательства', 'Лицо, в пользу которого установлено ограничение прав и обременение объекта недвижимости',
                       'Наличие незасвидетельствованных данных', 'Прочие ограничения ЗУ', 'Тип', 'Кадастровые номера расположенных в пределах земельного участка объектов недвижимости/Кадастровые номера иных объектов недвижимости, в пределах которых расположен объект недвижимости',
                       'Кдастровая стоимость']

            for r, row in enumerate(list):
                for c, col in enumerate(row):
                    worksheet.write(0, c, _header_name[c])
                    worksheet.write(r+1, c, col)
            workbook.close()
            QMessageBox.information(self,"Success!",f"Данные сохранены в файле: \n{self.fileName[0]}")

    def back_menu(self):
        MW = MainWindow()
        widget.addWidget(MW)
        widget.setCurrentIndex(widget.currentIndex() + 1)

    def re_name(self):
        self.file_name = QFileDialog.getOpenFileNames(self, 'Открыть файл', '', "*.xml")
        self.file_name_next = self.file_name[0]

        if not self.file_name_next:
            return
        else:

            self.list = []
            model = self.table.model()
            for _row in range(model.rowCount()):
                self.list.append("{}".format(model.index(_row, 4).data() or ""))

            for files in range(len(self.file_name_next)):
                if self.file_name_next[files] == (path.dirname(self.file_name_next[files])+"/"+self.list[files].replace(":","_")+".pdf"):
                    next
                else:
                    rename(self.file_name_next[files], str(path.dirname(self.file_name_next[files])+"/"+self.list[files].replace(":","_")+".xml"))

    def clear_table(self):
        self.table.setRowCount(0)

class MainWeb(QMainWindow):
#     # Переопределяем конструктор класса
    def __init__(self):
#         # Обязательно нужно вызвать метод супер класса
        super().__init__()
        self.setMinimumSize(QSize(1400, 600))  # Устанавливаем размеры
        self.setWindowTitle("Получение информации из Web")  # Устанавливаем заголовок окна

        central_widget = QWidget(self) # Создаём центральный виджет
        self.setCentralWidget(central_widget)
        self.createtoolbar()
        self.createMenuBar()
        layout = QGridLayout()
        central_widget.setLayout(layout)

        self.table = QTableWidget()  # Создаём таблицу
        self.table.setGeometry(10, 10, 1880, 900)

        self.table.setColumnCount(12)  # Устанавливаем три колонки

        self.table.setHorizontalHeaderLabels(["№", "Кадастровый номер", "       Адрес       ", "Тип", "Площадь", "Этаж", "Форма собственности",
                                              "Для постановки на учёт", "Кадастровая стоимость", "Дата определения стоимости",
                                                "Дата внесения стоимости в базу", "Категория земели"])

        self.table.resizeColumnsToContents()
        layout.addWidget(self.table, 0, 0)


    def createtoolbar(self):
        getAction_1 = QAction("Загрузить кад. номера", self)
        getAction_1.triggered.connect(self.get_file_name)

        getAction_2 = QAction("Получить информацию", self)
        getAction_2.triggered.connect(self.full_table)

        getAction_3 = QAction("Выгрузить в Excel", self)
        getAction_3.triggered.connect(self.fileSave)

        getAction_4 = QAction("Очистить таблицу", self)
        getAction_4.triggered.connect(self.clear_table)

        self.toolbar = self.addToolBar('Загрузить')
        self.toolbar.addAction(getAction_1)
        self.toolbar.addAction(getAction_2)
        self.toolbar.addAction(getAction_3)
        self.toolbar.addAction(getAction_4)


    def createMenuBar(self):
        self.menuBar = self.menuBar()
        fileMenu = self.menuBar.addMenu("Меню")
        back_action = QAction("Назад", self)
        back_action.triggered.connect(self.back_menu)
        fileMenu.addAction(back_action)

    def get_file_name(self):
        self.file_name = QFileDialog.getOpenFileNames(self, 'Открыть файл', '', "*.xls, *.xlsx")

        if not self.file_name[0]:
            return
        else:
            wb = openpyxl.load_workbook(self.file_name[0][0])
            ws = wb.active

            self.cell_value = []
            for i in range(2, ws.max_row+1):
                self.cell_value.append(ws.cell(column=1, row=i).value)

            self.table.setRowCount(len(self.cell_value))
            numb_row = len(self.cell_value)
            n = 0
            for row in range(numb_row):
                n += 1
                self.table.setItem(row, 0, QTableWidgetItem(str(n)))
                self.table.setItem(row, 1, QTableWidgetItem(str(self.cell_value[row])))

            self.table.resizeColumnsToContents()
            self.table.resizeRowsToContents()

    def full_table(self):
        list = []
        model = self.table.model()
        for _row in range(model.rowCount()):
            list.append("{}".format(model.index(_row, 1).data() or ""))

        self.list_next = []
        for count in range(len(list)):
            self.list_next.append(get_data_from_page(list[count]))

        for row in range(model.rowCount()):
            for column in range(2,12):
                self.table.setItem(row, column, QTableWidgetItem(str(self.list_next[row][column-2])))

        self.table.resizeRowsToContents()

    def fileSave(self):
        self.fileName = QFileDialog.getSaveFileName(self, "Сохранить файл", ".", "All Files(*.xlsx)")

        if not self.fileName[0]:
            return
        else:

            list = []
            model = self.table.model()
            for _row in range(model.rowCount()):
                r = []
                for _column in range(model.columnCount()):
                    r.append("{}".format(model.index(_row, _column).data() or ""))
                list.append(r)


            workbook = Workbook(self.fileName[0])
            worksheet = workbook.add_worksheet()

            _header_name = ["№", "Кадастровый номер", "       Адрес       ", "Тип", "Площадь", "Этаж", "Форма собственности",
                                              "Для постановки на учёт", "Кадастровая стоимость", "Дата определения стоимости",
                                                "Дата внесения стоимости в базу", "Категория земели"]

            for r, row in enumerate(list):
                for c, col in enumerate(row):
                    worksheet.write(0, c, _header_name[c])
                    worksheet.write(r+1, c, col)
            workbook.close()
            QMessageBox.information(self,"Success!",f"Данные сохранены в файле: \n{self.fileName[0]}")

    def back_menu(self):
        MW = MainWindow()
        widget.addWidget(MW)
        widget.setCurrentIndex(widget.currentIndex() + 1)

    def clear_table(self):
        self.table.setRowCount(0)


if __name__ == "__main__":
    app = QApplication(sys.argv)
    widget = QStackedWidget()
    MW = MainWindow()
    widget.addWidget(MW)
    widget.show()
    sys.exit(app.exec())


